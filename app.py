import streamlit as st
import pandas as pd
from datetime import datetime, time
import gspread
from google.oauth2.service_account import Credentials
import io

# ─── CONFIG ───────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Relación de Desplazamiento | ARL Bolívar",
    page_icon="🗺️",
    layout="wide"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.main { background-color: #F7F8FA; }
.header-bar {
    background: linear-gradient(135deg, #003087 0%, #0052CC 100%);
    padding: 2rem 2.5rem; border-radius: 12px;
    margin-bottom: 2rem; color: white;
}
.header-bar h1 { margin: 0; font-size: 1.8rem; font-weight: 700; }
.header-bar p  { margin: 0.3rem 0 0; font-size: 0.95rem; opacity: 0.85; }
.login-card {
    background: white; border-radius: 12px; padding: 2.5rem;
    max-width: 420px; margin: 4rem auto;
    border: 1px solid #E5E9F0;
    box-shadow: 0 4px 20px rgba(0,0,0,0.08);
}
.login-title { text-align:center; font-size:1.3rem; font-weight:700; color:#003087; margin-bottom:1.5rem; }
.section-card {
    background: white; border-radius: 10px; padding: 1.5rem 2rem;
    margin-bottom: 1.5rem; border: 1px solid #E5E9F0;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
}
.section-title {
    font-size:1rem; font-weight:600; color:#003087;
    border-left:4px solid #0052CC; padding-left:0.75rem; margin-bottom:1.2rem;
}
.badge-registrador { display:inline-block; background:#E8F0FE; color:#1A56DB; padding:0.3rem 0.9rem; border-radius:20px; font-size:0.85rem; font-weight:600; margin-bottom:1rem; }
.badge-validador   { display:inline-block; background:#FEF3C7; color:#92400E; padding:0.3rem 0.9rem; border-radius:20px; font-size:0.85rem; font-weight:600; margin-bottom:1rem; }
.val-aprobado { background:#E6F4EA; border:1px solid #34A853; border-radius:8px; padding:0.75rem 1rem; color:#1E6E3A; font-weight:500; margin:0.3rem 0; }
.val-nocumple { background:#FDE8E8; border:1px solid #EA4335; border-radius:8px; padding:0.75rem 1rem; color:#9B1C1C; font-weight:500; margin:0.3rem 0; }
.stButton > button {
    background: linear-gradient(135deg, #003087, #0052CC);
    color: white; border: none; padding: 0.75rem 2.5rem;
    border-radius: 8px; font-weight: 600; font-size: 1rem; width: 100%;
}
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTES ───────────────────────────────────────────────────────────────
SHEET_ID   = "1dr2kyYtuTOZeFmrA5QhzdAq0BCMQ4VR7lfXNY0m0WeY"
TOLERANCIA = 500
SCOPES     = ["https://www.googleapis.com/auth/spreadsheets",
              "https://www.googleapis.com/auth/drive"]

USUARIOS = {
    "123456789": ("123456789", "registrador", "Registrador"),
    "validador":  ("validador2026", "validador",   "Validador General"),
}

COLUMNAS = [
    "Marca temporal", "Registrado por", "Documento de identidad",
    "Codigo Sipab", "Nombre PGR", "Fecha", "Nombre AGR",
    "Empresa Cliente", "Cronograma", "Secuencia", "Empresa",
    "Recorrido", "Origen", "Destino", "Hora Inicio", "Hora Fin",
    "Frecuencia", "Valor Pasajes", "Transporte Interno", "Desayuno",
    "Almuerzo/Cena", "Hospedaje", "Detalles de las rutas", "Total",
    "Validacion tarifa", "Tarifa permitida", "Detalle validacion",
    "Estado", "Aprueba/No aprueba", "Observacion validador",
    "Validado por", "Fecha validacion",
]

FRECUENCIAS = ["Única visita","Diaria","Semanal","Quincenal","Mensual"]
RECORRIDOS  = ["Ida","Ida y vuelta"]

# ─── GOOGLE SHEETS ────────────────────────────────────────────────────────────
@st.cache_resource
def conectar_sheets():
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"], scopes=SCOPES
    )
    return gspread.authorize(creds)

def obtener_hoja():
    gc = conectar_sheets()
    sh = gc.open("Registros Desplazamiento")
    try:
        ws = sh.worksheet("Registros")
    except:
        ws = sh.add_worksheet(title="Registros", rows=1000, cols=40)
        ws.append_row(COLUMNAS)
    return ws

def cargar_registros():
    try:
        ws  = obtener_hoja()
        data = ws.get_all_records()
        return pd.DataFrame(data) if data else pd.DataFrame(columns=COLUMNAS)
    except Exception as e:
        st.error(f"Error cargando datos: {e}")
        return pd.DataFrame(columns=COLUMNAS)

def guardar_registro(datos: dict):
    ws  = obtener_hoja()
    fila = [str(datos.get(col, "")) for col in COLUMNAS]
    ws.append_row(fila)

def actualizar_decision(row_num, decision, observacion, validador):
    ws = obtener_hoja()
    # row_num es índice 0-based del df → fila en sheet = row_num + 2 (1 header + 1 offset)
    fila_sheet = row_num + 2
    col_apr  = COLUMNAS.index("Aprueba/No aprueba") + 1
    col_obs  = COLUMNAS.index("Observacion validador") + 1
    col_val  = COLUMNAS.index("Validado por") + 1
    col_fval = COLUMNAS.index("Fecha validacion") + 1
    ws.update_cell(fila_sheet, col_apr,  decision)
    ws.update_cell(fila_sheet, col_obs,  observacion)
    ws.update_cell(fila_sheet, col_val,  validador)
    ws.update_cell(fila_sheet, col_fval, str(datetime.now()))

# ─── TARIFAS ──────────────────────────────────────────────────────────────────
@st.cache_data
def cargar_tarifas():
    try:
        wb = pd.read_excel("DESPLAZAMIENTOS_2026.xlsx",
                           sheet_name="DESPLAZAMIENTOS 2026",
                           skiprows=1, engine="openpyxl")
        wb.columns = wb.columns.str.strip().str.upper()
        wb = wb.rename(columns={wb.columns[0]:"ORIGEN", wb.columns[1]:"DESTINO",
                                  wb.columns[2]:"VALOR_IDA", wb.columns[3]:"VALOR_IDA_VUELTA"})
        wb = wb[["ORIGEN","DESTINO","VALOR_IDA","VALOR_IDA_VUELTA"]].dropna(subset=["ORIGEN","DESTINO"])
        wb["ORIGEN"]  = wb["ORIGEN"].astype(str).str.strip().str.upper()
        wb["DESTINO"] = wb["DESTINO"].astype(str).str.strip().str.upper()
        ciudades = sorted(set(wb["ORIGEN"].tolist() + wb["DESTINO"].tolist()))
        return wb, ciudades
    except:
        return pd.DataFrame(), ["BOGOTÁ","MEDELLÍN","CALI","BARRANQUILLA","BUCARAMANGA",
                                 "CARTAGENA","VILLAVICENCIO","YOPAL","ACACÍAS"]

df_desp, CIUDADES = cargar_tarifas()

def validar_tarifa(origen, destino, recorrido, val_pasajes, transporte_interno):
    if df_desp.empty:
        return "SIN DATOS", None, "Tabla de tarifas no cargada"
    o = origen.strip().upper()
    d = destino.strip().upper()
    fila = df_desp[(df_desp["ORIGEN"]==o) & (df_desp["DESTINO"]==d)]
    if fila.empty:
        fila = df_desp[(df_desp["ORIGEN"]==d) & (df_desp["DESTINO"]==o)]
    if fila.empty:
        return "NO CUMPLE", None, "Ruta no encontrada en tabla de tarifas"
    fila = fila.iloc[0]
    tarifa = fila["VALOR_IDA_VUELTA"] if "ida y vuelta" in recorrido.lower() else fila["VALOR_IDA"]
    try:
        tarifa = float(tarifa)
    except:
        return "NO CUMPLE", None, "Tarifa no disponible"
    valor = val_pasajes + transporte_interno
    if valor <= tarifa + TOLERANCIA:
        return "APROBADO", tarifa, f"Valor ${valor:,.0f} ≤ Tarifa ${tarifa:,.0f}"
    else:
        return "NO CUMPLE", tarifa, f"Valor ${valor:,.0f} supera tarifa ${tarifa:,.0f} en ${valor-tarifa:,.0f}"

# ─── PLANTILLA ────────────────────────────────────────────────────────────────
def generar_plantilla():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PLANTILLA DESPLAZAMIENTO"

    campos = [
        "Documento de identidad","Codigo Sipab","Nombre PGR",
        "Fecha (DD/MM/AAAA)","Nombre AGR","Empresa Cliente",
        "Cronograma","Secuencia","Empresa de transporte",
        "Recorrido (Ida / Ida y vuelta)","Origen","Destino",
        "Hora Inicio (HH:MM)","Hora Fin (HH:MM)","Frecuencia",
        "Valor Pasajes Intermunicipal","Transporte Interno",
        "Desayuno","Almuerzo/Cena","Hospedaje","Detalles de las rutas",
    ]
    ejemplos = [
        "1012345678","7643","JUAN PEREZ","04/06/2026",
        "MARIA GOMEZ","BOLIVAR S.A.","100001","1",
        "EXPRESO BOGOTA","Ida y vuelta","BOGOTA","MEDELLIN",
        "06:00","10:00","Unica visita",
        "150000","30000","10000","15000","0",
        "Ruta por autopista principal",
    ]

    hf  = PatternFill(start_color="003087", end_color="003087", fill_type="solid")
    ef  = PatternFill(start_color="EBF2FF", end_color="EBF2FF", fill_type="solid")
    brd = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"),  bottom=Side(style="thin"))

    for i, campo in enumerate(campos, 1):
        c = ws.cell(row=1, column=i, value=campo)
        c.fill = hf; c.font = Font(color="FFFFFF", bold=True, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
        ws.column_dimensions[get_column_letter(i)].width = max(len(campo)+4, 20)

    for i, ej in enumerate(ejemplos, 1):
        c = ws.cell(row=2, column=i, value=ej)
        c.fill = ef; c.font = Font(color="555555", italic=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 25
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─── SESIÓN ───────────────────────────────────────────────────────────────────
if "autenticado" not in st.session_state:
    st.session_state.autenticado    = False
    st.session_state.rol            = None
    st.session_state.nombre_usuario = None
    st.session_state.usuario_id     = None

# ─── LOGIN ────────────────────────────────────────────────────────────────────
def pantalla_login():
    st.markdown("""
    <div class="header-bar">
        <h1>🗺️ Relación de Desplazamiento</h1>
        <p>ARL Bolívar · Outsourcing Adecco</p>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="login-card">', unsafe_allow_html=True)
    st.markdown('<div class="login-title">🔐 Iniciar Sesión</div>', unsafe_allow_html=True)
    usuario  = st.text_input("Usuario (Cédula)", placeholder="Ej: 123456789")
    password = st.text_input("Contraseña", type="password")
    if st.button("Ingresar"):
        if usuario in USUARIOS:
            pwd, rol, nombre = USUARIOS[usuario]
            if password == pwd:
                st.session_state.autenticado    = True
                st.session_state.rol            = rol
                st.session_state.nombre_usuario = nombre
                st.session_state.usuario_id     = usuario
                st.rerun()
            else:
                st.error("❌ Contraseña incorrecta")
        else:
            st.error("❌ Usuario no encontrado")
    st.markdown('</div>', unsafe_allow_html=True)

# ─── REGISTRADOR ──────────────────────────────────────────────────────────────
def vista_registrador():
    st.markdown("""
    <div class="header-bar">
        <h1>🗺️ Relación de Desplazamiento</h1>
        <p>ARL Bolívar · Outsourcing Adecco — Registro de desplazamientos</p>
    </div>""", unsafe_allow_html=True)

    col_b, col_l = st.columns([8,1])
    with col_b:
        st.markdown(f'<span class="badge-registrador">👤 {st.session_state.nombre_usuario}</span>', unsafe_allow_html=True)
    with col_l:
        if st.button("Salir"): st.session_state.autenticado = False; st.rerun()

    tab_form, tab_masivo, tab_mis = st.tabs(["📝 Nuevo Registro","📤 Cargue Masivo","📋 Mis Registros"])

    # ── TAB FORMULARIO ────────────────────────────────────────────────────────
    with tab_form:
        with st.form("form_desplazamiento", clear_on_submit=True):

            st.markdown('<div class="section-card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title">👤 Datos del PGR</div>', unsafe_allow_html=True)
            c1,c2,c3 = st.columns(3)
            with c1: documento    = st.text_input("Documento de Identidad")
            with c2: codigo_sipab = st.text_input("Código SIPAB")
            with c3: nombre_pgr   = st.text_input("Nombre PGR")
            c4,c5,c6 = st.columns(3)
            with c4: fecha           = st.date_input("Fecha", value=datetime.today())
            with c5: nombre_agr      = st.text_input("Nombre AGR")
            with c6: empresa_cliente = st.text_input("Empresa Cliente")
            c7,c8 = st.columns(2)
            with c7: cronograma = st.text_input("Cronograma")
            with c8: secuencia  = st.text_input("Secuencia")
            st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('<div class="section-card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title">🚗 Datos del Desplazamiento</div>', unsafe_allow_html=True)
            c9,c10 = st.columns(2)
            with c9:  empresa   = st.text_input("Empresa de transporte")
            with c10: recorrido = st.selectbox("Recorrido", RECORRIDOS)
            c11,c12 = st.columns(2)
            with c11: origen  = st.selectbox("Origen",  ["Seleccione..."] + CIUDADES)
            with c12: destino = st.selectbox("Destino", ["Seleccione..."] + CIUDADES)
            c13,c14,c15 = st.columns(3)
            with c13: hora_inicio = st.time_input("Hora Inicio", value=time(6,0))
            with c14: hora_fin    = st.time_input("Hora Fin",    value=time(10,0))
            with c15: frecuencia  = st.selectbox("Frecuencia", FRECUENCIAS)
            st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('<div class="section-card">', unsafe_allow_html=True)
            st.markdown('<div class="section-title">💰 Valores</div>', unsafe_allow_html=True)
            c16,c17,c18,c19 = st.columns(4)
            with c16: val_pasajes        = st.number_input("Pasajes ($)",          min_value=0, step=1000, value=0, format="%d")
            with c17: transporte_interno = st.number_input("Transporte Interno ($)", min_value=0, step=1000, value=0, format="%d")
            with c18: desayuno_val       = st.number_input("Desayuno ($)",         min_value=0, step=1000, value=0, format="%d")
            with c19: almuerzo_cena_val  = st.number_input("Almuerzo/Cena ($)",    min_value=0, step=1000, value=0, format="%d")
            c20,c21 = st.columns([1,3])
            with c20: hospedaje_val  = st.number_input("Hospedaje ($)", min_value=0, step=1000, value=0, format="%d")
            with c21: detalles_rutas = st.text_area("Detalles de las Rutas", height=80)
            total = val_pasajes + transporte_interno + desayuno_val + almuerzo_cena_val + hospedaje_val
            st.info(f"💵 **Total estimado: ${total:,.0f}**")
            st.markdown('</div>', unsafe_allow_html=True)

            submitted = st.form_submit_button("✅ Enviar Registro")

        if submitted:
            if origen == "Seleccione..." or destino == "Seleccione...":
                st.error("⚠️ Selecciona Origen y Destino.")
            elif not nombre_pgr or not documento:
                st.error("⚠️ Documento y Nombre PGR son obligatorios.")
            else:
                res3, tarifa, msg3 = validar_tarifa(origen, destino, recorrido, val_pasajes, transporte_interno)
                datos = {
                    "Marca temporal":           str(datetime.now()),
                    "Registrado por":           st.session_state.usuario_id,
                    "Documento de identidad":   documento,
                    "Codigo Sipab":             codigo_sipab,
                    "Nombre PGR":               nombre_pgr,
                    "Fecha":                    str(fecha),
                    "Nombre AGR":               nombre_agr,
                    "Empresa Cliente":          empresa_cliente,
                    "Cronograma":               cronograma,
                    "Secuencia":                secuencia,
                    "Empresa":                  empresa,
                    "Recorrido":                recorrido,
                    "Origen":                   origen,
                    "Destino":                  destino,
                    "Hora Inicio":              str(hora_inicio),
                    "Hora Fin":                 str(hora_fin),
                    "Frecuencia":               frecuencia,
                    "Valor Pasajes":            val_pasajes,
                    "Transporte Interno":       transporte_interno,
                    "Desayuno":                 desayuno_val,
                    "Almuerzo/Cena":            almuerzo_cena_val,
                    "Hospedaje":                hospedaje_val,
                    "Detalles de las rutas":    detalles_rutas,
                    "Total":                    total,
                    "Validacion tarifa":        res3,
                    "Tarifa permitida":         str(tarifa) if tarifa else "",
                    "Detalle validacion":       msg3,
                    "Estado":                   "PENDIENTE",
                    "Aprueba/No aprueba":       "",
                    "Observacion validador":    "",
                    "Validado por":             "",
                    "Fecha validacion":         "",
                }
                guardar_registro(datos)
                css = "val-aprobado" if res3 == "APROBADO" else "val-nocumple"
                st.markdown(f'<div class="{css}">🔍 <b>Validación tarifa:</b> {res3} — {msg3}</div>', unsafe_allow_html=True)
                st.success("✅ Registro enviado. Pendiente de aprobación por el validador.")

    # ── TAB CARGUE MASIVO ─────────────────────────────────────────────────────
    with tab_masivo:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">📤 Cargue Masivo de Registros</div>', unsafe_allow_html=True)

        st.markdown("**Paso 1 — Descarga la plantilla y diligénciala.**")
        st.download_button(
            label="⬇️ Descargar Plantilla Excel",
            data=generar_plantilla(),
            file_name="plantilla_desplazamiento.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_plantilla"
        )

        st.markdown("---")
        st.markdown("**Paso 2 — Sube el archivo diligenciado.**")
        archivo_masivo = st.file_uploader("Selecciona el archivo Excel", type=["xlsx","xls"])

        if archivo_masivo is not None:
            st.markdown("**Paso 3 — Haz clic en Subir para guardar los registros.**")
            if st.button("📤 Subir Registros"):
                try:
                    df_m = pd.read_excel(io.BytesIO(archivo_masivo.read()), engine="openpyxl")
                    df_m.columns = df_m.columns.str.strip()
                    df_m = df_m[df_m.iloc[:,0].astype(str).str.strip() != "1012345678"]
                    df_m = df_m.dropna(how="all")

                    errores = []
                    guardados = 0

                    for i, fila in df_m.iterrows():
                        doc = str(fila.get("Documento de identidad","")).strip()
                        pgr = str(fila.get("Nombre PGR","")).strip()
                        if not doc or doc=="nan" or not pgr or pgr=="nan":
                            errores.append(f"Fila {i+2}: Documento o Nombre PGR vacío")
                            continue

                        orig = str(fila.get("Origen","")).strip().upper()
                        dest = str(fila.get("Destino","")).strip().upper()
                        rec  = str(fila.get("Recorrido (Ida / Ida y vuelta)","")).strip()

                        try:
                            vp  = float(str(fila.get("Valor Pasajes Intermunicipal",0) or 0).replace(",","").replace("$",""))
                            ti  = float(str(fila.get("Transporte Interno",0) or 0).replace(",","").replace("$",""))
                            des = float(str(fila.get("Desayuno",0) or 0).replace(",","").replace("$",""))
                            alm = float(str(fila.get("Almuerzo/Cena",0) or 0).replace(",","").replace("$",""))
                            hos = float(str(fila.get("Hospedaje",0) or 0).replace(",","").replace("$",""))
                        except:
                            errores.append(f"Fila {i+2}: Error en valores numéricos")
                            continue

                        total_fila = vp + ti + des + alm + hos
                        res3, tarifa, msg3 = validar_tarifa(orig, dest, rec, vp, ti)

                        datos = {
                            "Marca temporal":           str(datetime.now()),
                            "Registrado por":           st.session_state.usuario_id,
                            "Documento de identidad":   doc,
                            "Codigo Sipab":             str(fila.get("Codigo Sipab","")).strip(),
                            "Nombre PGR":               pgr,
                            "Fecha":                    str(fila.get("Fecha (DD/MM/AAAA)","")),
                            "Nombre AGR":               str(fila.get("Nombre AGR","")).strip(),
                            "Empresa Cliente":          str(fila.get("Empresa Cliente","")).strip(),
                            "Cronograma":               str(fila.get("Cronograma","")).strip(),
                            "Secuencia":                str(fila.get("Secuencia","")).strip(),
                            "Empresa":                  str(fila.get("Empresa de transporte","")).strip(),
                            "Recorrido":                rec,
                            "Origen":                   orig,
                            "Destino":                  dest,
                            "Hora Inicio":              str(fila.get("Hora Inicio (HH:MM)","")),
                            "Hora Fin":                 str(fila.get("Hora Fin (HH:MM)","")),
                            "Frecuencia":               str(fila.get("Frecuencia","")).strip(),
                            "Valor Pasajes":            vp,
                            "Transporte Interno":       ti,
                            "Desayuno":                 des,
                            "Almuerzo/Cena":            alm,
                            "Hospedaje":                hos,
                            "Detalles de las rutas":    str(fila.get("Detalles de las rutas","")).strip(),
                            "Total":                    total_fila,
                            "Validacion tarifa":        res3,
                            "Tarifa permitida":         str(tarifa) if tarifa else "",
                            "Detalle validacion":       msg3,
                            "Estado":                   "PENDIENTE",
                            "Aprueba/No aprueba":       "",
                            "Observacion validador":    "",
                            "Validado por":             "",
                            "Fecha validacion":         "",
                        }
                        guardar_registro(datos)
                        guardados += 1

                    if guardados > 0:
                        st.success(f"✅ {guardados} registro(s) cargados. Quedan pendientes de aprobación.")
                    if errores:
                        st.warning("⚠️ Filas con errores:")
                        for e in errores: st.write(f"- {e}")
                except Exception as e:
                    st.error(f"❌ Error procesando el archivo: {e}")

        st.markdown('</div>', unsafe_allow_html=True)

    # ── TAB MIS REGISTROS ─────────────────────────────────────────────────────
    with tab_mis:
        df = cargar_registros()
        if df.empty:
            st.info("📭 Aún no tienes registros.")
        else:
            mis = df[df["Registrado por"].astype(str) == st.session_state.usuario_id].copy() if "Registrado por" in df.columns else df.copy()
            st.markdown(f"**{len(mis)} registro(s)**")

            if "Aprueba/No aprueba" in mis.columns:
                opciones = ["Todos"] + sorted(mis["Aprueba/No aprueba"].dropna().unique().tolist())
                filtro = st.selectbox("Filtrar por estado", opciones)
                if filtro != "Todos":
                    mis = mis[mis["Aprueba/No aprueba"] == filtro]

            st.dataframe(mis, use_container_width=True, hide_index=True)

            buf = io.BytesIO()
            mis.to_excel(buf, index=False, engine="openpyxl")
            st.download_button("⬇️ Descargar mis registros", buf.getvalue(),
                file_name="mis_registros.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── VALIDADOR ────────────────────────────────────────────────────────────────
def vista_validador():
    st.markdown("""
    <div class="header-bar">
        <h1>🗺️ Relación de Desplazamiento</h1>
        <p>ARL Bolívar · Outsourcing Adecco — Panel de Validación</p>
    </div>""", unsafe_allow_html=True)

    col_b, col_l = st.columns([8,1])
    with col_b:
        st.markdown(f'<span class="badge-validador">🔍 {st.session_state.nombre_usuario}</span>', unsafe_allow_html=True)
    with col_l:
        if st.button("Salir"): st.session_state.autenticado = False; st.rerun()

    tab_pend, tab_hist = st.tabs(["⏳ Pendientes","📋 Historial Completo"])
    df = cargar_registros()

    with tab_pend:
        if df.empty:
            st.info("📭 No hay registros aún.")
        else:
            pendientes = df[df["Aprueba/No aprueba"].astype(str).str.strip() == ""].copy()
            st.markdown(f"**{len(pendientes)} registro(s) pendiente(s)**")

            if pendientes.empty:
                st.success("✅ No hay registros pendientes.")
            else:
                for idx, row in pendientes.iterrows():
                    titulo = f"📄 {row.get('Nombre PGR','—')} | {row.get('Origen','—')} → {row.get('Destino','—')} | ${str(row.get('Total',0))} | {row.get('Fecha','')}"
                    with st.expander(titulo):
                        c1,c2,c3 = st.columns(3)
                        with c1:
                            st.markdown(f"**Documento:** {row.get('Documento de identidad','—')}")
                            st.markdown(f"**SIPAB:** {row.get('Codigo Sipab','—')}")
                            st.markdown(f"**AGR:** {row.get('Nombre AGR','—')}")
                        with c2:
                            st.markdown(f"**Empresa Cliente:** {row.get('Empresa Cliente','—')}")
                            st.markdown(f"**Cronograma:** {row.get('Cronograma','—')}")
                            st.markdown(f"**Recorrido:** {row.get('Recorrido','—')}")
                        with c3:
                            st.markdown(f"**Pasajes:** ${row.get('Valor Pasajes',0)}")
                            st.markdown(f"**Transporte:** ${row.get('Transporte Interno',0)}")
                            st.markdown(f"**Total:** **${row.get('Total',0)}**")

                        res3 = row.get("Validacion tarifa","—")
                        css  = "val-aprobado" if res3 == "APROBADO" else "val-nocumple"
                        st.markdown(f'<div class="{css}">🔍 <b>Validación tarifa:</b> {res3} — {row.get("Detalle validacion","")}</div>', unsafe_allow_html=True)

                        st.markdown("---")
                        obs = st.text_area("Observación", key=f"obs_{idx}", placeholder="Escribe una observación...")
                        bc1,bc2 = st.columns(2)
                        with bc1:
                            if st.button("✅ Aprobar", key=f"apr_{idx}"):
                                actualizar_decision(idx, "APROBADO", obs, st.session_state.nombre_usuario)
                                st.success("✅ Aprobado.")
                                st.rerun()
                        with bc2:
                            if st.button("❌ Rechazar", key=f"rec_{idx}"):
                                actualizar_decision(idx, "RECHAZADO", obs, st.session_state.nombre_usuario)
                                st.error("❌ Rechazado.")
                                st.rerun()

    with tab_hist:
        if df.empty:
            st.info("📭 No hay registros.")
        else:
            st.markdown(f"**{len(df)} registro(s) en total**")

            cf1,cf2 = st.columns(2)
            with cf1:
                pgr_f = st.multiselect("Filtrar por PGR",
                    options=sorted(df["Nombre PGR"].dropna().unique()) if "Nombre PGR" in df.columns else [])
            with cf2:
                est_f = st.multiselect("Filtrar por Estado",
                    options=sorted(df["Aprueba/No aprueba"].dropna().unique()) if "Aprueba/No aprueba" in df.columns else [])

            df_f = df.copy()
            if pgr_f: df_f = df_f[df_f["Nombre PGR"].isin(pgr_f)]
            if est_f: df_f = df_f[df_f["Aprueba/No aprueba"].isin(est_f)]

            st.dataframe(df_f, use_container_width=True, hide_index=True)

            buf = io.BytesIO()
            df_f.to_excel(buf, index=False, engine="openpyxl")
            st.download_button("⬇️ Descargar Excel", buf.getvalue(),
                file_name="historial_desplazamientos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─── ROUTER ───────────────────────────────────────────────────────────────────
if not st.session_state.autenticado:
    pantalla_login()
elif st.session_state.rol == "registrador":
    vista_registrador()
elif st.session_state.rol == "validador":
    vista_validador()
